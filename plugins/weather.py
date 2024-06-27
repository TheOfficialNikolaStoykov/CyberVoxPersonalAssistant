import requests
from openpyxl import load_workbook
from synthesize_speech import VoiceSynthesizer


class Weather:

    def __init__(self):
        self.city_name = None
        self.city_list = []
        self.final_answer = None
        # Instantiate the VoiceSynthesizer class
        self.synthesizer = VoiceSynthesizer('')


    def generate_city_list(self):
        # Populate the self.city_list with the 1-st column in the cities_list.xlsx file
        workbook = load_workbook('./cities_list.xlsx')

        sheet = workbook['Sheet1']

        for row in sheet.iter_rows(min_row=2, max_row=372, min_col=1, max_col=1, values_only=True):
            self.city_list.append(row[0])

        workbook.close()


    def keyword_parser(self, command, phrase):
        # Parse the phrase, check if the city in the phrase is in the self.city_list
        self.generate_city_list()
        tokens = phrase.split()
        for city in self.city_list:
            for token in tokens:
                if city == token:
                    # Set the self.city_name to the city in the phrase
                    self.city_name = city
        self.execute()


    def execute(self):
        # The main function that executes the weather plugin
        self.generate_url()
        self.return_weather_data()


    def generate_url(self):
        # Generate an URL for the OpenWeatherMap API
        BASE_URL = 'https://api.openweathermap.org/data/2.5/weather?'
        API_KEY = '4ab20ee5a669e09e73903fee178be875'

        url = BASE_URL + 'appid=' + API_KEY + '&q=' + self.city_name

        return url
    

    def kelvin_to_celc(self, kelvin):
        # Convert Kelvin to Celcius
        celcius = kelvin - 273.15

        return celcius
    

    def return_weather_data(self):
        # Return the weather data
        response = requests.get(self.generate_url()).json()
        description = response['weather'][0]['description']
        temp_kelvin = response['main']['temp']
        temp_celc = int(self.kelvin_to_celc(temp_kelvin))
        temp_kelvin_min = response['main']['temp_min']
        temp_celc_min = int(self.kelvin_to_celc(temp_kelvin_min))
        temp_kelvin_max = response['main']['temp_max']
        temp_celc_max = int(self.kelvin_to_celc(temp_kelvin_max))
        humidity = response['main']['humidity']

        self.final_answer = f'The temperature in {self.city_name} is {temp_celc}degrees. The minimum temperature is {temp_celc_min}degrees and the maximum temperature is {temp_celc_max}degrees. Humidity is {humidity}%. Description is {description}.'
        
        # Save the self.final_answer as output.mp3 and speak it
        self.synthesizer.synthesize_speech(self.final_answer)